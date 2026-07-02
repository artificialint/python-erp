"""Fixed-template import pipeline: read .xlsx -> validate -> upsert into DB.

Invalid rows are recorded in ``import_errors`` and NOT imported; every run writes an
``import_batches`` summary. Idempotent: masters upsert by business code. Each row runs
in a SAVEPOINT so one bad row cannot poison the others.

See docs/PYTHON_ERP_INV0_PLAN.md §5.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from erp_data.db import models
from erp_data.db.base import (
    DocumentType,
    ImportOutcome,
    ItemType,
    OrgType,
    PartyType,
    RuleType,
)
from erp_data.imports.templates import TEMPLATES


class RowError(Exception):
    """A row-level validation/resolution failure targeted at an optional column."""

    def __init__(self, code: str, message: str, column: str | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.column = column


# ── coercion helpers ─────────────────────────────────────────────────
def _s(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _float(value: Any, column: str) -> float | None:
    v = _s(value)
    if v is None:
        return None
    try:
        return float(v)
    except ValueError as exc:
        raise RowError("invalid_number", f"{column} must be a number", column) from exc


def _bool(value: Any) -> bool:
    v = _s(value)
    return v.lower() in {"1", "true", "yes", "y", "evet", "x"} if v else False


def _date(value: Any, column: str) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _s(value)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise RowError("invalid_date", f"{column} must be a date (YYYY-MM-DD)", column)


def _enum(value: Any, enum_cls: Any, column: str) -> Any:
    v = _s(value)
    if v is None:
        return None
    try:
        return enum_cls(v)
    except ValueError as exc:
        allowed = ", ".join(e.value for e in enum_cls)
        raise RowError("invalid_value", f"{column} must be one of: {allowed}", column) from exc


# ── reference resolution ─────────────────────────────────────────────
def _resolve_org(session: Session, code: str | None, column: str) -> models.Organization | None:
    if not code:
        return None
    org = session.execute(
        select(models.Organization).where(models.Organization.code == code)
    ).scalar_one_or_none()
    if org is None:
        raise RowError("unresolved_reference", f"organization code '{code}' not found", column)
    return org


def _resolve_product(session: Session, code: str | None, column: str) -> models.ProductService | None:
    if not code:
        return None
    product = session.execute(
        select(models.ProductService).where(models.ProductService.code == code)
    ).scalar_one_or_none()
    if product is None:
        raise RowError("unresolved_reference", f"product code '{code}' not found", column)
    return product


# ── per-template upsert handlers ─────────────────────────────────────
def _h_companies(session: Session, row: dict) -> None:
    code = _s(row.get("code"))
    org = session.execute(
        select(models.Organization).where(models.Organization.code == code)
    ).scalar_one_or_none()
    if org is None:
        org = models.Organization(code=code, legal_name=_s(row.get("legal_name")))
        session.add(org)
    else:
        org.legal_name = _s(row.get("legal_name")) or org.legal_name
    org.org_type = _enum(row.get("org_type"), OrgType, "org_type") or OrgType.company
    for field in (
        "address", "city", "country", "phone", "email", "tax_no",
        "default_currency", "bank_name", "bank_account", "iban", "swift_code",
    ):
        setattr(org, field, _s(row.get(field)))
    parent_code = _s(row.get("parent_code"))
    if parent_code:
        session.flush()
        org.parent_id = _resolve_org(session, parent_code, "parent_code").id


def _h_customers(session: Session, row: dict) -> None:
    name = _s(row.get("company_name"))
    party = session.execute(
        select(models.Party).where(
            models.Party.party_type == PartyType.customer,
            models.Party.legal_name == name,
        )
    ).scalar_one_or_none()
    if party is None:
        party = models.Party(party_type=PartyType.customer, legal_name=name)
        session.add(party)
    for field in ("address", "city", "country", "phone", "email", "tax_no"):
        setattr(party, field, _s(row.get(field)))
    party.default_discount_percent = _float(row.get("default_discount_percent"), "default_discount_percent")
    party.default_payment_term_code = _s(row.get("default_payment_term_code"))


def _h_products(session: Session, row: dict) -> None:
    code = _s(row.get("product_code"))
    product = session.execute(
        select(models.ProductService).where(models.ProductService.code == code)
    ).scalar_one_or_none()
    if product is None:
        product = models.ProductService(code=code)
        session.add(product)
    product.description = _s(row.get("product_description"))
    product.hs_code = _s(row.get("hs_code"))
    product.unit = _s(row.get("unit"))
    product.item_type = _enum(row.get("item_type"), ItemType, "item_type") or ItemType.product
    product.product_tax_percent = _float(row.get("product_tax_percent"), "product_tax_percent")
    org = _resolve_org(session, _s(row.get("organization_code")), "organization_code")
    product.organization_id = org.id if org else None


def _h_payment_terms(session: Session, row: dict) -> None:
    code = _s(row.get("code"))
    term = session.execute(
        select(models.PaymentTerm).where(models.PaymentTerm.code == code)
    ).scalar_one_or_none()
    if term is None:
        term = models.PaymentTerm(code=code, label=_s(row.get("label")))
        session.add(term)
    term.label = _s(row.get("label")) or term.label
    term.is_default = _bool(row.get("is_default"))


def _h_docno_rules(session: Session, row: dict) -> None:
    org = _resolve_org(session, _s(row.get("organization_code")), "organization_code")
    doc_type = _enum(row.get("document_type"), DocumentType, "document_type")
    rule = session.execute(
        select(models.DocnoRule).where(
            models.DocnoRule.organization_id == org.id,
            models.DocnoRule.document_type == doc_type,
        )
    ).scalar_one_or_none()
    if rule is None:
        rule = models.DocnoRule(organization_id=org.id, document_type=doc_type)
        session.add(rule)
    rule.template = _s(row.get("template"))
    rule.counter_scope = _s(row.get("counter_scope"))


def _h_tax_rules(session: Session, row: dict) -> None:
    product = _resolve_product(session, _s(row.get("product_code")), "product_code")
    org = _resolve_org(session, _s(row.get("organization_code")), "organization_code")
    session.add(
        models.TaxRule(
            rule_type=_enum(row.get("rule_type"), RuleType, "rule_type"),
            seller_country=_s(row.get("seller_country")),
            buyer_country=_s(row.get("buyer_country")),
            product_id=product.id if product else None,
            organization_id=org.id if org else None,
            rate=_float(row.get("rate"), "rate") or 0.0,
            reason=_s(row.get("reason")),
        )
    )


def _h_price_lists(session: Session, row: dict) -> None:
    code = _s(row.get("price_list_code"))
    price_list = session.execute(
        select(models.PriceList).where(models.PriceList.code == code)
    ).scalar_one_or_none()
    if price_list is None:
        org = _resolve_org(session, _s(row.get("organization_code")), "organization_code")
        price_list = models.PriceList(
            code=code,
            organization_id=org.id if org else None,
            currency=_s(row.get("currency")),
            valid_from=_date(row.get("valid_from"), "valid_from"),
            valid_to=_date(row.get("valid_to"), "valid_to"),
        )
        session.add(price_list)
        session.flush()
    product = _resolve_product(session, _s(row.get("product_code")), "product_code")
    if product is None:
        raise RowError("required_field_missing", "product_code is required", "product_code")
    item = session.execute(
        select(models.PriceListItem).where(
            models.PriceListItem.price_list_id == price_list.id,
            models.PriceListItem.product_id == product.id,
        )
    ).scalar_one_or_none()
    if item is None:
        item = models.PriceListItem(price_list_id=price_list.id, product_id=product.id, unit_price=0.0)
        session.add(item)
    item.unit_price = _float(row.get("unit_price"), "unit_price") or 0.0
    item.min_quantity = _float(row.get("min_quantity"), "min_quantity")


def _h_employees(session: Session, row: dict) -> None:
    name = _s(row.get("person_name"))
    party = session.execute(
        select(models.Party).where(
            models.Party.party_type == PartyType.person,
            models.Party.legal_name == name,
        )
    ).scalar_one_or_none()
    if party is None:
        party = models.Party(party_type=PartyType.person, legal_name=name)
        session.add(party)
    party.email = _s(row.get("email"))
    party.phone = _s(row.get("phone"))
    party.tax_no = _s(row.get("tax_no"))


HANDLERS: dict[str, Callable[[Session, dict], None]] = {
    "companies": _h_companies,
    "customers": _h_customers,
    "products": _h_products,
    "payment_terms": _h_payment_terms,
    "docno_rules": _h_docno_rules,
    "tax_rules": _h_tax_rules,
    "price_lists": _h_price_lists,
    "employees": _h_employees,
}


def read_rows(path: Path | str) -> tuple[list[str], list[dict]]:
    """Return (headers, data-rows-as-dicts); fully-blank rows are skipped."""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook.active
    raw_rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not raw_rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
    data: list[dict] = []
    for raw in raw_rows[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        data.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
    return headers, data


def _record_error(session: Session, batch_id: int, row_number: int, error: RowError) -> None:
    session.add(
        models.ImportRowError(
            import_batch_id=batch_id,
            row_number=row_number,
            column_name=error.column,
            error_code=error.code,
            error_message=error.message,
        )
    )
    session.flush()


def import_file(
    session: Session,
    template_type: str,
    path: Path | str,
    *,
    created_by_employee_id: int | None = None,
) -> models.ImportBatch:
    """Import a fixed-template .xlsx into the DB; return the ImportBatch summary."""
    spec = TEMPLATES[template_type]
    handler = HANDLERS[template_type]
    headers, data_rows = read_rows(path)

    batch = models.ImportBatch(
        template_type=template_type,
        file_name=Path(path).name,
        created_by_employee_id=created_by_employee_id,
    )
    session.add(batch)
    session.flush()

    missing_cols = [col for col in spec.columns if col not in headers]
    if missing_cols:
        _record_error(
            session, batch.id, 1,
            RowError("missing_columns", f"Template columns missing: {', '.join(missing_cols)}",
                     ", ".join(missing_cols)),
        )
        batch.total_rows = len(data_rows)
        batch.valid_rows = 0
        batch.invalid_rows = len(data_rows)
        batch.outcome = ImportOutcome.failed
        session.commit()
        return batch

    valid = invalid = 0
    for row_number, row in enumerate(data_rows, start=2):  # row 1 = header
        missing_req = [col for col in spec.required if not _s(row.get(col))]
        if missing_req:
            _record_error(
                session, batch.id, row_number,
                RowError("required_field_missing", f"Missing required: {', '.join(missing_req)}", missing_req[0]),
            )
            invalid += 1
            continue
        savepoint = session.begin_nested()
        try:
            handler(session, row)
            session.flush()
            savepoint.commit()
            valid += 1
        except RowError as err:
            savepoint.rollback()
            _record_error(session, batch.id, row_number, err)
            invalid += 1
        except Exception as err:  # noqa: BLE001 — defensive per-row boundary
            savepoint.rollback()
            _record_error(session, batch.id, row_number, RowError("row_failed", str(err)))
            invalid += 1

    batch.total_rows = len(data_rows)
    batch.valid_rows = valid
    batch.invalid_rows = invalid
    batch.outcome = ImportOutcome.completed if invalid == 0 else ImportOutcome.completed_with_errors
    session.commit()
    return batch
